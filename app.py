import requests
from bs4 import BeautifulSoup
import sqlite3
import openpyxl
import datetime
from datetime import datetime, time, timedelta
import pandas as pd
import re
import customtkinter as ctk
from tkinter import filedialog
from tkinter import messagebox
import os


def extract_numeric_values(data):
    numeric_values = []
    for row in data:
        numeric_row = []
        for value in row:
            # Extract numeric part from the value
            numeric_value = re.sub(r'[^\d.-]', '', value)
            numeric_row.append(numeric_value)
        numeric_values.append(numeric_row)
    return numeric_values

def fetch_html_content(url):
    # Send a GET request to the URL
    response = requests.get(url)
    
    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content
        soup = BeautifulSoup(response.content, 'html.parser')
        return soup
    else:
        raise Exception(f"Failed to fetch page content: {response.status_code}")

def extract_weather_data(soup):
    # Locate the relevant elements containing the weather data
    data_elements = soup.find_all('tr', class_='ng-star-inserted')
    
    # Extract the headers
    headers = ['Time', 'Temperature', 'Dew_Point', 'Humidity', 'Wind', 'Speed', 'Gust', 'Pressure', 'Precip_Rate', 'Precip_Accum', 'UV', 'Solar']
    print("Headers:", headers)
    
    # Extract the rows
    rows = []
    for element in data_elements:
        cells = element.find_all('td')
        
        if len(cells) > 1:  # Ensure there are enough cells in the row
            row_data = [cell.get_text(strip=True) for cell in cells]
            rows.append(row_data)
    
    print("Rows:", rows[0])
    return headers, rows

def create_database():
    # Create the database in the user's home directory
    home_dir = os.path.expanduser("~")
    db_dir = os.path.join(home_dir, "WeatherData")
    os.makedirs(db_dir, exist_ok=True)  # Create the directory if it does not exist
    db_path = os.path.join(db_dir, 'weather_data.db')
    conn = sqlite3.connect(db_path)
    
    c = conn.cursor()
    
    c.execute('''
        CREATE TABLE IF NOT EXISTS observations (
            Time TEXT,
            Temperature TEXT,
            Dew_Point TEXT,
            Humidity TEXT,
            Wind TEXT,
            Speed TEXT,
            Gust TEXT,
            Pressure TEXT,
            Precip_Rate TEXT,
            Precip_Accum TEXT,
            UV TEXT,
            Solar TEXT
        )
    ''')
    conn.commit()
    c.execute('DELETE FROM observations')

    # Print the column names
    c.execute('PRAGMA table_info(observations)')
    columns = [info[1] for info in c.fetchall()]
    print("Database columns:", columns)
    return conn

def insert_observations(conn, headers, rows):
    c = conn.cursor()
    for row in rows:
        row[0] = datetime.strptime(row[0], '%I:%M %p').strftime('%H:%M')
        print("Inserting row:", row)
        c.execute(f'''
            INSERT INTO observations ({", ".join(headers)}) VALUES ({", ".join(["?"] * len(headers))})
        ''', row)
    conn.commit()

def query_sorted_data(conn):
    c = conn.cursor()
    c.execute('SELECT * FROM observations ORDER BY Time')
    rows = c.fetchall()
    return rows

def extract_45_min_intervals(conn, start_time_str):
    c = conn.cursor()
    # Parse the start time
    start_time = datetime.strptime(start_time_str, '%H:%M')
    # Query all observations
    c.execute('SELECT * FROM observations ORDER BY Time')
    rows = c.fetchall()
    
    if not rows:
        print("No data in the database.")
        return []
    
    filtered_data = []
    current_time = start_time
    row_index = 0  # To track the current row being processed

    while current_time < (start_time + timedelta(hours=24)) and row_index < len(rows):
        # Find the closest match to the current_time
        closest_row = None
        min_diff = timedelta.max

        for row in rows[row_index:]:
            row_time = datetime.strptime(row[0], '%H:%M')
            time_diff = abs(row_time - current_time)
            if time_diff < min_diff:
                min_diff = time_diff
                closest_row = row
                row_index += 1
            else:
                break  # If the difference increases, we've passed the closest match
        
        if closest_row:
            filtered_data.append(closest_row)
        current_time += timedelta(minutes=45)  # Increment by 45 minutes
    
    return filtered_data

def save_to_spreadsheet(filtered_data, headers, filename='filtered_weather_data.xlsx'):
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Weather Data'

    # Write headers
    ws.append(headers)
    # Write filtered data
    for row in filtered_data:
        ws.append(row)

    # Save the workbook
    wb.save(filename)
    print(f"Filtered data saved to {filename}")

def save_all_data_to_spreadsheet(conn, headers, filename='Erie2.xlsx', start_row=7, start_col=2):
    # Load the existing workbook
    wb = openpyxl.load_workbook(filename)
    ws = wb['Field Data Entry']  # Select the active worksheet or specify the sheet name with ws = wb['SheetName']

    # Get the initial time from cell B7
    time_cell = ws.cell(row=start_row, column=start_col).value
    print(time_cell)

    # Convert the time format from the Excel sheet to the expected format
    if isinstance(time_cell, datetime):
        time_str = time_cell.strftime('%H:%M')
    elif isinstance(time_cell, time):
        time_str = time_cell.strftime('%H:%M')
    else:
        time_str = str(time_cell)
        if len(time_str) == 4 and time_str.isdigit():
            time_str = f"{time_str[:2]}:{time_str[2:]}"
        else:
            try:
                time_str = datetime.strptime(time_str, '%I:%M:%S %p').strftime('%H:%M')
            except ValueError:
                time_str = datetime.strptime(time_str, '%H:%M').strftime('%H:%M')

    print(time_str)

    # Query the data for 45-minute intervals starting from the initial time
    filtered_data = extract_45_min_intervals(conn, time_str)

    # Define the specific columns for each data field
    column_mapping = {
        'Time': 'B',
        'Temperature': 'BU',
        'Wind': 'BT',  # direction
        'Speed': 'BQ',  # wind speed
        'Gust': 'BR',  # wind gust
        'Precip_Rate': 'BP',
        # Add other mappings as needed
    }

    # Write data to specific cells every 45 rows
    for i, row in enumerate(filtered_data):
        row_num = start_row + (i * 15)  # Increment by 15 rows
        speed_value = None
        gust_value = None
        for field, col_letter in column_mapping.items():
            col_num = openpyxl.utils.column_index_from_string(col_letter)
            cell = ws.cell(row=row_num, column=col_num)
            if field in ['Speed', 'Gust', 'Temperature'] and cell.value is None:
                numeric_value = re.sub(r'[^\d.-]', '', row[headers.index(field)])
                numeric_value = float(numeric_value) if numeric_value else None
                cell.value = (numeric_value)//1
                if field == 'Speed':
                    speed_value = (numeric_value)
                    print(f"Speed: {speed_value}")
                elif field == 'Gust':
                    gust_value = (numeric_value)
                    print(f"Gust: {gust_value}")
            elif field == "Wind" and cell.value is None:
                wind_direction = row[headers.index(field)]
                if wind_direction == "N":
                    cell.value = 0
                elif wind_direction == "NNE":
                    cell.value = 22
                elif wind_direction == "NE":
                    cell.value = 45
                elif wind_direction == "ENE":
                    cell.value = 67
                elif wind_direction == "E":
                    cell.value = 90
                elif wind_direction == "ESE":
                    cell.value = 112
                elif wind_direction == "SE":
                    cell.value = 135
                elif wind_direction == "SSE":
                    cell.value = 157
                elif wind_direction == "S":
                    cell.value = 180
                elif wind_direction == "SSW":
                    cell.value = 202
                elif wind_direction == "SW":
                    cell.value = 225
                elif wind_direction == "WSW":
                    cell.value = 247
                elif wind_direction == "W":
                    cell.value = 270
                elif wind_direction == "WNW":
                    cell.value = 292
                elif wind_direction == "NW":
                    cell.value = 315
                elif wind_direction == "NNW":
                    cell.value = 337
                else:
                    cell.value = None  # Handle unexpected values
            elif field == "Precip_Rate":
                if cell.value == None:
                    rain_numeric_value = re.sub(r'[^\d.-]', '', row[headers.index(field)])
                    rain_numeric_value = float(rain_numeric_value) if rain_numeric_value else 0.0
                    print(f"Rain numeric value: {rain_numeric_value}")
                    if rain_numeric_value == 0.0:
                        cell.value = "No Rain"
                    elif rain_numeric_value <= 0.1:
                        cell.value = "Light Rain"
                    else:
                        cell.value = "Heavy Rain"
            else:
                continue 
                #cell.value = row[headers.index(field)]

        # Calculate the value for column BS based on Speed and Gust
        if speed_value is not None and gust_value is not None:
            print(f"Calculating BS value for Speed: {speed_value} and Gust: {gust_value}")
            if speed_value < 15:
                bs_value = speed_value//1
            else:
                bs_value = (speed_value * 2 / 3 + gust_value * 1 / 3)//1

            bs_col_num = openpyxl.utils.column_index_from_string('BS')
            bs_cell = ws.cell(row=row_num, column=bs_col_num)
            bs_cell.value = bs_value
    # Save the workbook
    wb.save(filename)
    print(f"All data saved to {filename}")

def main():
    def run_script():
        try:
            url = url_entry.get().strip().strip('"')
            filename = filename_entry.get().strip()
            
            # Fetch and parse the HTML content
            soup = fetch_html_content(url)

            # Extract the weather data from the soup
            headers, rows = extract_weather_data(soup)

            # Create the database and insert the observations
            conn = create_database()
            insert_observations(conn, headers, rows)

            # Query and print the sorted data
            sorted_data = query_sorted_data(conn)

            print(extract_45_min_intervals(conn, "06:00"))

            save_all_data_to_spreadsheet(conn, headers, filename=filename, start_row=7, start_col=2)

            # Show success message
            messagebox.showinfo("Success", "Data extraction and saving completed successfully!")

        except Exception as e:
            # Show error message
            messagebox.showerror("Error", f"An error occurred: {e}")

    def browse_file():
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            filename_entry.delete(0, ctk.END)
            filename_entry.insert(0, filename)

    app = ctk.CTk()
    app.title("Weather Data Extractor")

    ctk.CTkLabel(app, text="Enter URL:").pack(pady=5)
    url_entry = ctk.CTkEntry(app, width=400)
    url_entry.pack(pady=5)

    ctk.CTkLabel(app, text="Select Excel File:").pack(pady=5)
    filename_entry = ctk.CTkEntry(app, width=400)
    filename_entry.pack(pady=5)
    ctk.CTkButton(app, text="Browse", command=browse_file).pack(pady=5)

    ctk.CTkButton(app, text="Run", command=run_script).pack(pady=20)

    app.mainloop()


if __name__ == "__main__":
    main()
